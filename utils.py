import openpyxl
import re
import os
import tests_classes


def natural_sort(list_to_sort):
    """Function to naturally sort lists containing strings with numbers and digits."""

    def alphabet_num_key(element):
        list_to_return = []
        for substring in re.split('([0-9]+)', element):
            if substring.isdigit():
                list_to_return.append(int(substring))
            else:
                list_to_return.append(substring.lower())
        return list_to_return

    return sorted(list_to_sort, key=alphabet_num_key)


def clean_out_path(out_path):
    """Initialize out_path, and removes older versions of copied assignments that were saved there"""

    # deleting old copies of the assignments in out_path
    if os.path.exists(out_path):
        for file in os.listdir(out_path):
            if os.path.isfile(out_path + file):
                if '.txt' not in file:  # some tests require having text files in the same directory as the assignments.
                    os.remove(os.path.join(out_path, file))
    else:
        os.mkdir(out_path)

    # Write a new file to save out_path as module. This enable to import files from out_path
    f = open(out_path + '__init__.py', 'w+')
    f.close()


def write_clean_code(input_file, out_file):
    """Copy original input_file, clean code outside blocks, and write it to out_file.
    Both files should be open."""
    middle_of_block = False
    for i, line in enumerate(input_file):
        # print(i)
        line = line.replace('\t', '    ')
        if line == '' or line == '\n':
            continue
        elif middle_of_block:
            if line[:1] != ' ':
                if line[:4] == 'def ' or line[:6] == 'class ':
                    out_file.write('\n\n')
                    out_file.write(line)
                    middle_of_block = True
                else:
                    middle_of_block = False
            else:
                if 'def ' in line or 'class ' in line:
                    out_file.write('\n')
                out_file.write(line)
        else:
            if line[:4] == 'def ' or line[:6] == 'class ':
                out_file.write('\n\n')
                out_file.write(line)
                middle_of_block = True
            elif line[:1] == '#' or line[:7] == 'import ' or line[:5] == 'from ':
                out_file.write(line)


def copy_clean(path, file_name, out_path):
    """Copies the assignments to out_path, cleans and deletes code outside blocks."""

    # opens the input file to copy from it.
    input_file = open(path + file_name, "r")

    # In case the assignment name includes '.' it should be changed in order to import from it.
    file_name = file_name[:-3].replace('.', '_') + file_name[-3:]
    file_to_test = open(out_path + file_name, "w")

    # Avoid the usage of "input" built-in function that stops the test from running.
    # Also adds the student id number as global variable in the file.
    begin_string = """def input(*args):
    return ''\n\n
FILE_NAME = '""" + str(file_name[:-3]) + """'\n\n"""

    file_to_test.write(begin_string)

    write_clean_code(input_file, file_to_test)

    # closes and save
    input_file.close()
    file_to_test.close()


def copy_assignments(path, out_path):
    """Rewrites assignments to new files in out_path, and cleans unwanted code blocks."""

    for j, file in enumerate(natural_sort(os.listdir(path))):
        if file[-3:] != '.py' or '__init__' in file:
            continue
        copy_clean(path, file, out_path)


def create_workbook(workbook_path, sheet_title):
    """Creates xlsx file to write to it all the comments and grades."""

    grades = openpyxl.Workbook()
    worksheet = grades.active
    worksheet.title = sheet_title
    worksheet['A1'] = 'Group ID'
    worksheet['B1'] = 'Grade'
    worksheet['C1'] = 'Grader Notes'
    grades.save(workbook_path)


def read_tests(test_path):
    """Reads tests file and returns it's data."""

    tests_file = open(test_path, "r")
    tests, points, expected, test_types = [], [], [], []
    for line in tests_file.readlines()[1:]:
        if len(line) > 1:
            params = line.strip("\n").split("|")
            tests.append(params[0])
            points.append(params[1])
            expected.append(params[2])
            test_types.append(params[3])

    return tests, points, expected, test_types


def run_tests(assignment_num, module_name, tests, points, expected, test_types, testing_stuck):
    """Testing function.
    For each test in tests it creates an instance of Test class, that fits the test type.
    It runs each test, saves the notes and sums the score"""
    score = 0
    notes = []

    for i in range(len(tests)):
        current_test = tests[i]
        current_points = points[i]
        current_expected = expected[i]
        current_test_type = test_types[i]

        # checks what kind of Test class fits the current test.
        if current_test_type in tests_classes.TEST_TYPES_DICT:
            # if the test type has dedicated class to it, make a new instance of it
            test_class = tests_classes.TEST_TYPES_DICT[current_test_type]
            test = test_class(assignment_num, current_test, current_points, current_expected, module_name,
                              testing_stuck)
        else:
            # else, make a regular Test class instance
            test = tests_classes.Test(assignment_num, current_test, current_points, current_expected, module_name,
                                      testing_stuck)

        # runs the test and saves the results
        current_score, current_note = test.compile_run_and_compare()

        # sums the score
        score += current_score

        # Adds notes from the last test to notes list.
        # Also removes unfamiliar calls in the notes to make it more accessible to students reading the notes
        if current_note != '':
            notes.append(current_note.replace('self.module.', ''))

    # Removes duplicates in notes, for instance compilation error.
    notes = list(dict.fromkeys(notes))

    return score, notes


def write_score_notes_to_file(score, notes, j, group_id, workbook_path):
    """Opens the workbook in the first sheet, and writes to the j-th row: group_id, score and notes."""

    # Opens workbook in the first sheet
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.worksheets[0]

    # writes group_id and score
    ws.cell(j, 1).value = group_id
    ws.cell(j, 2).value = score

    # writes all notes to one cell, but in it, every note in a new line.
    notes_str = ''
    if len(notes) == 0:
        notes_str = 'Good job! Keep it up! :)'
    else:
        for t in range(len(notes)):
            if t == 0:
                notes_str = notes[t]
            else:
                notes_str = notes_str + '\n' + notes[t]
    ws.cell(j, 3).value = notes_str

    # saves and closes the workbook
    wb.save(workbook_path)


def run_write(assignment_num, out_path, workbook_path, tests, points, expected, test_types, testing_stuck,
              flag_print=False):
    """Runs tests on each assignment and write its' score and notes to workbook"""
    for j, file_name in enumerate(natural_sort(os.listdir(out_path))):
        if file_name[-3:] != '.py' or '__init__' in file_name:
            continue

        module_name_to_import = out_path.replace('/', '.') + file_name[:-3]

        if assignment_num == 4:
            os.chdir(r'assignments/4/')
        score, notes = run_tests(assignment_num, module_name_to_import, tests, points, expected, test_types,
                                 testing_stuck)

        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        group_id = file_name[:file_name.find('A')]
        write_score_notes_to_file(score, notes, j+2, group_id, workbook_path)

        # print the file_name, its' score and notes
        if flag_print:
            print(file_name)
            print(score)
            for note in notes:
                print(note)
